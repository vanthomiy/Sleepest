from io import StringIO
from RecordDataObject import RecordDataObject
from ParamsObject import Parameter
import json
from openpyxl import load_workbook
import asyncio
from Algorithm import importNewSleepDataSet
from AdjustModel import adjustParamsRandom
from AdjustModel import historyCreater

#Lade die Excelliste mit allen verfügbaren Daten
def LoadExcelSheetsData():
  workbook = load_workbook(filename="Schlafdaten.xlsx")
  
  sheets = workbook.sheetnames
  
  sheetData = {}

  for sheetName in sheets:
    sheet = workbook[sheetName]
    maxRow = sheet.max_row
    for i in range(2,maxRow+1):

      datum=sheet.cell(row=i,column=1)
      time=sheet.cell(row=i,column=2)
      sleep=sheet.cell(row=i,column=3)
      light=sheet.cell(row=i,column=4)
      motion=sheet.cell(row=i,column=5)
      real=sheet.cell(row=i,column=6)

      record = RecordDataObject(datum.value, time.value, sleep.value, light.value, motion.value, real.value)
      sheetData.setdefault(sheetName, []).append(record)

  return sheetData



  #merge data of all


  return workbook._sheets



#Lade die Aktuellen und die Startparameter in das File
def LoadParams(values):
    params = []
    with open('Params.txt') as json_file:
         data = json.load(json_file)
         for dataObj in data[values]:
            params.append(Parameter(dataObj['art'], dataObj['min'], dataObj['max'], dataObj['set'], dataObj['offset'], dataObj['faktor']))
         return params


#Vergleiche richtig mit falsch und speichere es in der history mitsamt den parametern
def evaluateCalculation(calculation, personData):
    
    sleepWrong = 0 #the real sleep
    awakeWrong = 0 #the real awake
    sleepRight = 0
    awakeRight = 0

    #just easy compare, without sleepstates
    for i in range(len(calculation)):

        if(personData[i].realSleep > 0):
            w = 3

        if(calculation[i] < 0.5 and personData[i].realSleep == 0):
            awakeRight+=1
        elif (calculation[i] > 0.5 and personData[i].realSleep > 0):
            sleepRight+=1
        elif (calculation[i] < 0.5 and personData[i].realSleep > 0):
            sleepWrong+=1
        elif (calculation[i] > 0.5 and personData[i].realSleep == 0):
            awakeWrong+=1

    return awakeRight, sleepRight, awakeWrong, sleepWrong


#Program main
async def main():

    sheetData = LoadExcelSheetsData()
    loadedParams = LoadParams('actualParameter')

    

    actualParams = loadedParams
    bestParams = loadedParams
    bestAlloverAttempt = 0

    for attempt in range(1000):

        sleepRightAll = 0
        awakeRightAll = 0
        sleepWrongAll = 0
        awakeWrongAll = 0

        #for each persondatasets
        for person in sheetData.items():
            
            sleepRightPerson = 0
            awakeRightPerson = 0
            sleepWrongPerson = 0
            awakeWrongPerson= 0
        
            #for each sleepdataset
            for dataSet in person[1]:
                # can simply be awaited to wait until it is complete:
                task = asyncio.create_task(importNewSleepDataSet(dataSet, actualParams))
                calculationResult = await task
                eval = evaluateCalculation(calculationResult[0], calculationResult[1])
                

                awakeRightPerson += eval[0]
                sleepRightPerson += eval[1]
                awakeWrongPerson += eval[2]
                sleepWrongPerson += eval[3]

            #get the amount of wrong or right values for person
            sleepRightAll += sleepRightPerson
            awakeRightAll += awakeRightPerson
            sleepWrongAll += sleepWrongPerson
            awakeWrongAll += awakeWrongPerson

            amountPerson = sleepRightPerson + awakeRightPerson + sleepWrongPerson + awakeWrongPerson
            pecentagePerson = ((sleepRightPerson + awakeRightPerson) / amountPerson) * 100
            percentagePersonSleep = (sleepRightPerson / (sleepRightPerson + sleepWrongPerson + 0.000001)) * 100
            percentagePersonAwake = (awakeRightPerson / (awakeRightPerson + awakeWrongPerson + 0.000001)) * 100


            #print(person[0] + " Amount: %i" % amountPerson)
            #print(person[0] + " has allover %0.2f"  % pecentagePerson)
            #print(person[0] + " has sleep %0.2f"  % percentagePersonSleep)
            #print(person[0] + " has awake  %0.2f"  % percentagePersonAwake + "\n")

        #get the complete amount of wrong or right values
        amountAll = sleepRightAll + awakeRightAll + sleepWrongAll + awakeWrongAll
        pecentageAll = ((sleepRightAll + awakeRightAll) / amountAll) * 100
        percentageAllSleep = (sleepRightAll / (sleepRightAll + sleepWrongAll + 0.00001)) * 100
        percentageAllAwake = (awakeRightAll / (awakeRightAll + awakeWrongAll+ 0.00001)) * 100

        print("Versuch %i" % attempt + " Value: %0.2f" % pecentageAll)

        #Save if its better then before
        if(bestAlloverAttempt < pecentageAll):
            bestParams = actualParams
            historyCreater(amountAll,pecentageAll,percentageAllSleep,percentageAllAwake, actualParams)
            bestAlloverAttempt = pecentageAll
            print("-------------------> New Best: %0.2f" % pecentageAll)

        #adjust the parameters random
        actualParams = adjustParamsRandom(bestParams)




    #print("Amount: %i" % amountAll)
    #print("Allover %0.2f"  % pecentageAll)
    #print("Sleep %0.2f"  % percentageAllSleep)
    #print("Awake  %0.2f"  % percentageAllAwake + "\n")


   
asyncio.run(main())


